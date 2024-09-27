import pandas as pd
import numpy as np
import os
import re
import zipfile
from datetime import datetime
from openpyxl import load_workbook
import win32com.client as win32
import glob


# Функция для пересохранения файла с использованием win32 без удаления
def resave_excel_using_win32(input_file):
    try:
        excel = win32.Dispatch('Excel.Application')
        excel.DisplayAlerts = False  # Отключаем предупреждения
        wb = excel.Workbooks.Open(input_file)

        # Пересохраняем файл под тем же именем
        # FileFormat=51 означает сохранить как .xlsx
        wb.SaveAs(input_file, FileFormat=51)

        # Закрываем книгу и выходим из Excel
        wb.Close(SaveChanges=True)
        excel.Application.Quit()
        print(f"Файл успешно пересохранен как: {input_file}")
    except Exception as e:
        print(f"Ошибка при пересохранении файла с помощью Excel: {e}")
    finally:
        # Обязательно закрываем Excel, если что-то пошло не так
        if excel is not None:
            excel.Application.Quit()


# Получаем путь к текущей директории проекта
current_directory = os.getcwd()

# Ищем zip-файлы в домашней директории
home_directory = os.path.expanduser('~')
zip_files = glob.glob(os.path.join(
    home_directory, '**', '*.zip'), recursive=True)
zip_files.sort(key=os.path.getctime, reverse=True)
zip_files_last = zip_files[0]
print("Последний найденный архив:", zip_files_last)

# Распаковываем zip архив
with zipfile.ZipFile(zip_files_last, 'r') as zip_ref:
    zip_ref.extractall(current_directory)

sink_name_path = 'My RTN Far-End(Sink, Suhrob).xlsx'

# Поиск всех Excel-файлов в текущей директории
excel_files = glob.glob(os.path.join(current_directory, '*.xlsx'))
excel_files = [f for f in excel_files if 'History_Performance_Data' in f]
excel_files.sort(key=os.path.getctime, reverse=True)

# Пересохраняем файл без его удаления
resave_excel_using_win32(excel_files[0])

# Чтение Excel into DataFrame
try:
    rtn_df = pd.read_excel(
        excel_files[0], engine='openpyxl', skiprows=7, sheet_name='Sheet1')
    print(rtn_df)
except FileNotFoundError:
    print(f"Файл {excel_files[0]} не найден.")
except Exception as e:
    print(f"Ошибка при чтении Excel файла: {e}")

rtn_df['End Time'] = pd.to_datetime(rtn_df['End Time']).dt.date
rtn_df = rtn_df[
    (rtn_df['Performance Event'] == 'TSL_AVG(dbm)') | (
        rtn_df['Performance Event'] == 'RSL_AVG(dbm)')
]

rtn_df = pd.pivot_table(rtn_df, values='Value CUR', index='Monitored Object',
                        columns=['Performance Event'], aggfunc='mean').reset_index()

rtn_df = rtn_df[~rtn_df['TSL_AVG(dbm)'].between(-100, 0)]
rtn_df = rtn_df[~rtn_df['RSL_AVG(dbm)'].between(-100, -80)]
rtn_df = rtn_df[~rtn_df['RSL_AVG(dbm)'].between(-48, 0)]

rtn_df['RTN Site List'] = [
    [] for _ in range(len(rtn_df))
]

rtn_df['RTN LINK'] = pd.Series(dtype='object')

regexRTN = re.compile(
    r'^[A-Za-z]{2}\d{4}$|^[A-Za-z]{3}\d{3}$|^[A-Za-a]{4}\d{2}')

for i in rtn_df.index:
    rtn_df['RTN Site List'][i] = re.split(
        r'[-_.(): ]', rtn_df['Monitored Object'][i])
    rtn_df['RTN Site List'][i] = [
        j for j in rtn_df['RTN Site List'][i] if regexRTN.search(j)]

rtn_df['RTN Site List Str'] = rtn_df['RTN Site List'].apply(
    lambda x: '-'.join(x))
min_rsl_idx = rtn_df.groupby('RTN Site List Str')['TSL_AVG(dbm)'].idxmin()
rtn_df = rtn_df.loc[min_rsl_idx]
rtn_df = rtn_df.reset_index()


def checks_two_array(list1, list2):
    if isinstance(list1, list) and isinstance(list2, list):
        return list1[0] in list2[1:] and list2[0] in list[1:]
    return False


for i in rtn_df.index:
    if not pd.isna(rtn_df['RTN LINK'][i]):
        continue
    link_name = rtn_df['Monitored Object'][i]
    site_list = rtn_df['RTN Site List'][i]
    rtn_df['RTN LINK'][i] = link_name
    for j in range(i+1, len(rtn_df)):
        if not pd.isna(rtn_df['RTN LINK'][j]):
            continue
        if checks_two_array(site_list, rtn_df['RTN LINK'][j]):
            rtn_df['RTN LINK'][j] = link_name
        else:
            continue

rtn_df = rtn_df.drop(
    columns=['index', 'TSL_AVG(dbm)', 'RTN Site List', 'RTN Site List Str'], axis=1
)

part_for_drop = [
    '-ODU-1(RTNRF-1)-RTNRF:1', '-MXXI4B-1(IF)-RTNRF:1', '-DMD4-1(IF1)-RTNRF:1', '-DMD4-2(IF2)-RTNRF:1', '-MODU-2(RTNRF-2)-RTNRF:1', '-MODU-1(RTNRF-1)-RTNRF:2',
    '-MODU-2(RTNRF-2)-RTNRF:2', '-MODU-1(RTNRF-1)-RTNRF:1', '-DMD4-1(NM1500-NM1127_A)-RTNRF:1', '-DMD4-2(NM1500-NM1127_B)-RTNRF:1',
]

# for idx in rtn_df.index:
#     link_value = rtn_df.at[idx, 'RTN LINK'] # получил текущее значение строки .at faster then .loc
#     for part in part_for_drop:
#         link_value = link_value.replace(part, '') # удалил не нужную часть

#     rtn_df.at[idx, 'RTN LINK'] = link_name

# Итерация по строкам DataFrame с использованием индексов
for idx in rtn_df.index:
    link_value = rtn_df.at[idx, 'RTN LINK']  # Получаем текущее значение строки

    # Удаляем все ненужные части строки
    for part in part_for_drop:
        # Заменяем ненужные части строки
        link_value = link_value.replace(part, '')

    # Сохраняем обновленное значение обратно в DataFrame
    rtn_df.at[idx, 'RTN LINK'] = link_value


# Чтение Excel into DataFrame
sink_name = pd.read_excel(
    sink_name_path, engine='openpyxl', sheet_name='Лист1')

rtn_df = pd.merge(
    rtn_df, sink_name[['RTN LINK', 'SINK NE']], on='RTN LINK', how='inner'
).reset_index(drop=True)

with pd.ExcelWriter(excel_files[0], engine='openpyxl', mode='a') as writer:
    rtn_df.to_excel(writer, sheet_name='Report Page', index=False)


# directory_path = r"C:/Users/suhrob.yusubaxmedov/downloads"
# dest_path = r"C:\projects\MSS_REPORT"

# input_file_path = r'C:\projects\MSS_REPORT\RTN.xlsx'
# output_file_path = r'C:\projects\MSS_REPORT\RTN_resaved.xlsx'
# rtn_repot_path = r'C:\projects\MSS_REPORT\RTN_resaved.xlsx'

# def extractor():
#     # Регулярное выражение для поиска архивов с изменяющейся датой в названии
#     pattern_zip = re.compile(r'History_Performance_Data_(\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})\.zip')

#     zip_files_with_dates = []  # Список для найденных архивов и их дат

#     # Поиск всех архивов в указанной директории
#     for file in os.listdir(directory_path):
#         match = pattern_zip.match(file)
#         if match:
#             date_str = match.group(1)  # Парсинг даты из имени файла
#             date_obj = datetime.strptime(date_str, '%Y-%m-%d_%H-%M-%S')
#             zip_files_with_dates.append((file, date_obj))

#     # Если найдены архивы, сортируем их по дате
#     if zip_files_with_dates:
#         zip_files_with_dates.sort(key=lambda x: x[1], reverse=True)

#         latest_zip_file = zip_files_with_dates[0][0]  # Самый свежий и дикий Performance Data из загрузок :)
#         zip_path = os.path.join(directory_path, latest_zip_file)

#         # Распаковка архива
#         with zipfile.ZipFile(zip_path, 'r') as zip_ref:
#             zip_ref.extractall(dest_path)

#         # Поиск извлеченного файла
#         extracted_rtn_file = os.listdir(dest_path)
#         pattern_rtn_file = re.compile(r'History_Performance_Data_(\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})\.xlsx')

#         for file in extracted_rtn_file:
#             if pattern_rtn_file.match(file):
#                 old_file_path = os.path.join(dest_path, file)
#                 new_file_path = os.path.join(dest_path, 'RTN.xlsx')  # Новое имя для файла History Performance Data
#                 if os.path.exists(new_file_path):
#                     os.remove(new_file_path)
#                 os.rename(old_file_path, new_file_path)


# def resave_excel_using_win32(input_path, output_path):
#     try:
#         excel = win32.Dispatch('Excel.Application')
#         wb = excel.Workbooks.Open(input_path)
#         wb.SaveAs(output_path, FileFormat=51)  # FileFormat=51 означает сохранить как .xlsx
#         wb.Close()
#         excel.Application.Quit()
#         print(f"Файл успешно пересохранен как: {output_path}")

#         # Удаление исходного файла после пересохранения
#         if os.path.exists(input_path):
#             os.remove(input_path)
#             print(f"Файл {input_path} удален.")
#     except Exception as e:
#         print(f"Ошибка при пересохранении файла с помощью Excel: {e}")


# # Вызов функций
# extractor()
# resave_excel_using_win32(input_file_path, output_file_path)

# rtn_df = pd.read_excel(rtn_repot_path, skiprows=7, sheet_name='Sheet1')

# rtn_df['End Time'] = pd.to_datetime(rtn_df['End Time']).dt.date
# tsl_df = rtn_df[rtn_df['Performance Event'] == 'TSL_AVG(dbm)']
# rsl_df = rtn_df[rtn_df['Performance Event'] == 'RSL_AVG(dbm)']

# tsl_pivot = tsl_df.pivot_table(
#     index='Monitored Object', columns='End Time', values='Value CUR', aggfunc='mean'
# )
# tsl_pivot['Mean TSL'] = tsl_pivot.mean(axis=1)

# rsl_pivot = rsl_df.pivot_table(
#     index='Monitored Object', columns='End Time', values='Value CUR', aggfunc='mean'
# )
# rsl_pivot['Mean RSL'] = rsl_pivot.mean(axis=1)

# # Объединяем данные по 'Monitored Object'
# merged_df = pd.merge(tsl_pivot['Mean TSL'], rsl_pivot, on='Monitored Object', how='inner').reset_index()

# # Применяем фильтры
# merged_df = merged_df[~merged_df['Mean TSL'].between(-100, 0)]
# merged_df = merged_df[~merged_df['Mean RSL'].between(-100, -80)]
# merged_df = merged_df[~merged_df['Mean RSL'].between(-48, 0)]

# # Определяем столбцы для проверки и обрабатываем NaN значения
# rows_to_check = merged_df.columns.difference(['Monitored Object', 'Mean RSL'])
# merged_df[rows_to_check] = merged_df[rows_to_check].fillna(-9999)

# # Создаем маску для фильтрации строк
# mask = (merged_df[rows_to_check] > -48).sum(axis=1) >= 2

# # Фильтруем строки, чтобы сохранить 'Monitored Object' и 'Mean RSL'
# filtered_df = merged_df[~mask].copy()  # Явное создание копии

# # Восстанавливаем значения NaN вместо заполненных значений с использованием .loc
# filtered_df.loc[:, rows_to_check] = filtered_df.loc[:, rows_to_check].replace(-9999, np.nan)

# filtered_df = filtered_df.drop(columns='Mean TSL', axis=1)

# book = load_workbook(rtn_repot_path)
# if 'Sheet2' or 'Sheet3' in book.sheetnames:
#     del book['Sheet2']
#     del book['Sheet3']
# book.save(rtn_repot_path)

# part_for_drop = [
#     '-ODU-1(RTNRF-1)-RTNRF:1', '-MXXI4B-1(IF)-RTNRF:1', '-DMD4-1(IF1)-RTNRF:1', '-DMD4-2(IF2)-RTNRF:1', '-MODU-2(RTNRF-2)-RTNRF:1', '-MODU-1(RTNRF-1)-RTNRF:2',
#     '-MODU-2(RTNRF-2)-RTNRF:2', '-MODU-1(RTNRF-1)-RTNRF:1', '-DMD4-1(NM1500-NM1127_A)-RTNRF:1', '-DMD4-2(NM1500-NM1127_B)-RTNRF:1'
# ]
# # Сохраняем данные в Excel
# with pd.ExcelWriter(rtn_repot_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
#     tsl_pivot.to_excel(writer, sheet_name='TSL PIVOT')
#     rsl_pivot.to_excel(writer, sheet_name='RSL PIVOT')
#     filtered_df.to_excel(writer, sheet_name='RSL & TSL MEAN', index=False)

# print('Report Created Successfully.')
